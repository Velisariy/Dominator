import csv
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
    
reader = csv.reader(open("some.csv", "r", newline=''), dialect='excel')


wb = Workbook()
ws = wb.worksheets[0]
ws.title = "Sheet 1"

    
for row_index, row in enumerate(reader):
    res = [float(s.strip("[]")) for s in row[2].split(',')]
    res = round(sum(res) / len(res), 4)
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        if column_index == 1:
            pass
        elif column_index == 2:
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = res
        else:
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

wb.save("some.xls")